"""
This script analyzes cardiac cultures on cMEAs, returning their conduction velocity. 
The file name is currently specified in the code itself - future releases will enable the user to analyze files as command line arguments.

Example usage:
    $ python CV_dev.py
"""

print("Importing Libraries...\n")
import os, time, math, sys, traceback
import matplotlib.pyplot as plt 
import numpy as np
from openpyxl import Workbook
from hsl_lib.MCS_Objects import CV_Region_Selector
from hsl_lib.MCS_Functions import get_data_txt, plot_mea_waveforms, plot_cmea_waveforms, get_channels_to_compare, get_spike_time_differences, get_filename, get_filenames, log_error, save_workbook



def main():

    cmea_electrodes = [61, 53, 52, 41, 44, 54, 51, 42, 43, 31]
    cmea_distances = [0.001 * x for x in range(0, len(cmea_electrodes))] # Electrodes are 1mm apart

    cmea_positions = []

    channels_to_read = cmea_electrodes

    #full_file_path = get_filename()

    filenames = get_filenames()

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws['A1'] = "File"
    ws['B1'] = "CV (mean)"
    ws['C1'] = "Frequency (Hz)"

    ws_row = 2
    for full_file_path in filenames:
        try:    
            all_channels, analog_channels, sampling_rate = get_data(full_file_path, channels_to_read)



            plot_cmea_waveforms(all_channels, full_file_path)

            print("Please enter the two channels you would like to compare")
            channels_to_compare_input = [int(x) for x in raw_input("---->>>> ").split()]
            channels_to_compare = get_channels_to_compare(all_channels, channels_to_compare_input)
            stim_channel = analog_channels[0]

            cv_region_selector = CV_Region_Selector(channels_to_compare, stim_channel)
            start_time, end_time = cv_region_selector.start_time, cv_region_selector.end_time

            #conduction_intervals = 
            spike_time_differences = get_spike_time_differences(channels_to_compare, start_time, end_time)
            spike_frequency = (len(spike_time_differences) + 1) / (end_time - start_time)
            distance = cmea_distances[cmea_electrodes.index(channels_to_compare[0].channel_number)] - cmea_distances[cmea_electrodes.index(channels_to_compare[1].channel_number)]

            conduction_velocity_mean = distance / np.mean(spike_time_differences) 

            print("Conduction Velocity Mean: " + str(conduction_velocity_mean) + " m/s")
        
            ws.cell(row=ws_row, column=1, value = full_file_path)
            ws.cell(row=ws_row, column=2, value = conduction_velocity_mean)
            ws.cell(row=ws_row, column=3, value = spike_frequency)
            ws_row = ws_row + 1
        except:
            e = traceback.format_exc()
            log_error(full_file_path, e)

    
    save_workbook('CV', wb)
    
    print("Done!")

if __name__ == '__main__':
    
    main()