"""Generates QT Intervals!"""
from __future__ import print_function
import os
import time
import math
import sys
import traceback
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from hsl_lib.MCS_Objects import QT_Region_Selector
from hsl_lib.MCS_Functions import get_data_txt, plot_mea_waveforms, plot_cmea_waveforms, \
                                  get_channels_to_compare, get_spike_time_differences, \
                                  get_channels_to_analyze, get_filename, get_filenames, \
                                  log_error, get_data, save_workbook




def main():
    """The main function for QT_dev.py"""

    cmea_electrodes = [61, 53, 52, 41, 44, 54, 51, 42, 43, 31]
    cmea_distances = [0.001 * x for x in range(0, len(cmea_electrodes))]


    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws['A1'] = "File"
    ws['B1'] = "QT (mean)"
    for i in range(len(cmea_electrodes)):
        ws.cell(row=1, column=i+1, value=cmea_electrodes[i])

    ws_row = 2

    filenames = get_filenames()
    for full_file_path in filenames:
        try:

            all_channels, analog_channels, sampling_rate = get_data(full_file_path, \
                                                                    cmea_electrodes)

            plot_cmea_waveforms(all_channels, full_file_path)

            print("Please enter the channels you would like to analyze")
            channels_to_analyze_input = [int(x) for x in raw_input("---->>>> ").split()]
            channels_to_analyze = get_channels_to_analyze(all_channels, channels_to_analyze_input)

            ws.cell(row=ws_row, column=1, value=full_file_path)
            for channel in channels_to_analyze:
                region_selector = QT_Region_Selector(channel)
                qt_interval = (region_selector.qt_end_point - \
                              region_selector.qt_start_point) * (1000/sampling_rate)
                print("Channel Average QT Interval: " + str(qt_interval) + " ms")
                ws.cell(row=ws_row, column=cmea_electrodes.index(channel), value=qt_interval)
            ws_row = ws_row + 1
        except:
            error = traceback.format_exc()
            log_error(full_file_path, error)

    save_workbook('QT', wb)


if __name__ == "__main__":
    main()
    